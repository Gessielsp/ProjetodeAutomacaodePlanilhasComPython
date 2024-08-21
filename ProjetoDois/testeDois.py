import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill

def extrair_dados(caminho_arquivo):
    df = pd.read_excel(caminho_arquivo, sheet_name='CONSULTA LIBERAÇÕES')
    return df

def transformar_dados(df):
    meses_pt_br_para_en = {
        'JAN': 'Jan', 'FEV': 'Feb', 'MAR': 'Mar', 'ABR': 'Apr', 'MAI': 'May',
        'JUN': 'Jun', 'JUL': 'Jul', 'AGO': 'Aug', 'SET': 'Sep', 'OUT': 'Oct',
        'NOV': 'Nov', 'DEZ': 'Dec'
    }

    def converter_mes_para_en(data_str):
        mes_pt_br = data_str.split('/')[0].upper()
        ano = data_str.split('/')[1]
        mes_en = meses_pt_br_para_en.get(mes_pt_br, mes_pt_br)
        return f"{mes_en}/{ano}"

    df['Emissão - Mês Sigla Completa (MMM/AAAA)'] = df['Emissão - Mês Sigla Completa (MMM/AAAA)'].apply(converter_mes_para_en)
    df['Emissão - Mês Sigla Completa (MMM/AAAA)'] = pd.to_datetime(df['Emissão - Mês Sigla Completa (MMM/AAAA)'], format='%b/%Y')

    df = df.sort_values(by='Emissão - Mês Sigla Completa (MMM/AAAA)', ascending=True)
    
    df = df.rename(columns={'Emissão - Mês Sigla Completa (MMM/AAAA)': 'Emissão MÊs'})
    
    colunas_desejadas = [
        'Emissão MÊs',
        'PF - Ação Nome',
        'Emitente - UG Código',
        'Favorecido Doc. Número',
        'Favorecido Doc. Nome',
        'PF Número',
        'Doc - Observação Texto',
        'PF - Recurso Código',
        'PF - Fonte Recursos Código',
        'PF - Valor Linha Valor'
    ]

    for coluna in colunas_desejadas:
        if coluna not in df.columns:
            raise ValueError(f"Coluna '{coluna}' não encontrada no arquivo de entrada.")

    df = df[colunas_desejadas]

    df_liberacoes = df[df['PF - Ação Nome'].str.contains('LIBERACAO DE RECURSO FINANCEIRO', case=False, na=False)]
    df_transferencias = df[df['PF - Ação Nome'].str.contains('TRANSFERENCIA DE RECURSO FINANCEIRO', case=False, na=False)]

    return df_liberacoes, df_transferencias


def aplicar_formato_colunas(writer, sheet_name):
    workbook = writer.book
    worksheet = workbook[sheet_name]

    bold_font = Font(bold=True)
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for cell in worksheet[1]:  
        cell.value = cell.value.upper() 
        cell.font = bold_font  
        cell.fill = fill_green 

def aplicar_formato_negrito_e_data(writer, sheet_name, col):
    workbook = writer.book
    worksheet = workbook[sheet_name]

    bold_font = Font(bold=True)

    for cell in worksheet[col]:
        cell.font = bold_font

def carregar_dados(df_liberacoes, df_transferencias, caminho_arquivo_saida):
    with pd.ExcelWriter(caminho_arquivo_saida, engine='openpyxl') as writer:
        df_liberacoes.to_excel(writer, sheet_name='Liberações', index=False)
        df_transferencias.to_excel(writer, sheet_name='Transferência', index=False)

        aplicar_formato_colunas(writer, 'Liberações')
        aplicar_formato_colunas(writer, 'Transferência')

        aplicar_formato_negrito_e_data(writer, 'Liberações', 'A')
        aplicar_formato_negrito_e_data(writer, 'Transferência', 'A')

consulta_path = r"C:\Users\gessiel.passos\Documents\planilhas\CONSULTA LIBERAÇÕES.xlsx"
output_path = 'C:/Users/gessiel.passos/Documents/planilhas/Planilha_Atualizada.xlsx'

df_consulta = extrair_dados(consulta_path)
df_liberacoes, df_transferencias = transformar_dados(df_consulta)
carregar_dados(df_liberacoes, df_transferencias, output_path)

print("Processo ETL concluído com sucesso. Planilhas 'Liberações' e 'Transferência' atualizadas.")
